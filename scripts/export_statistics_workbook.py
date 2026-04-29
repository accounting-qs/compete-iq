"""
One-off conversion script: Excel workbook → JSON fixture for Statistics v1.

Usage:
    python scripts/export_statistics_workbook.py path/to/workbook.xlsx

Outputs api/data/statistics_workbook_snapshot.json — committed to the repo.
The app reads the JSON at runtime; no xlsx parsing at startup.
"""
from __future__ import annotations

import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# ---------------------------------------------------------------------------
# Column mapping: 1-based column index → semantic key
# Columns A (1) through BC (55).  BD–CK are scratch/unused.
# ---------------------------------------------------------------------------
COLUMN_MAP: dict[int, str] = {
    1: "webinarNumber",
    2: "dateOrNote",
    3: "status",
    4: "listUrl",
    5: "description",
    6: "listNameOrTitle",
    7: "sendInfo",
    8: "descLabel",
    9: "titleText",
    # 10–12 (listSize/listRemain/gcalInvited) removed from Statistics — skipped on export
    13: "accountsNeeded",
    14: "createdDate",
    15: "industry",
    16: "employeeRange",
    17: "country",
    18: "invited",
    19: "unsubscribes",
    # 20 (ghlPageViews) removed — not synced, intentionally skipped
    21: "lpRegs",
    22: "yesMarked",
    23: "yesAttended",
    24: "yes10MinPlus",
    25: "yesAttendBySmsClick",
    26: "yesBookings",
    27: "maybeMarked",
    28: "maybeAttended",
    29: "maybe10MinPlus",
    30: "maybeAttendBySmsClick",
    31: "maybeBookings",
    32: "selfRegMarked",
    33: "selfRegAttended",
    34: "selfReg10MinPlus",
    35: "selfRegBookings",
    36: "totalRegs",
    37: "totalAttended",
    38: "attendBySmsReminder",
    39: "total10MinPlus",
    40: "total30MinPlus",
    41: "totalBookings",
    42: "totalCallsDatePassed",
    43: "confirmed",
    44: "shows",
    45: "noShows",
    46: "canceled",
    47: "won",
    48: "disqualified",
    49: "qualified",
    50: "leadQualityGreat",
    51: "leadQualityOk",
    52: "leadQualityBarelyPassable",
    53: "leadQualityBadDq",
    54: "avgProjectedDealSize",
    55: "avgClosedDealValue",
}

# Metric columns (numeric data) — everything from col 10 onward
METRIC_KEYS = {k: v for k, v in COLUMN_MAP.items() if k >= 10}

# Text/identity columns
TEXT_KEYS = {k: v for k, v in COLUMN_MAP.items() if k < 10}

# Safe row bounds — exclude scratch/capacity notes after row 305
MAX_SAFE_ROW = 305
# Exception: webinar 136 starts at 301 and has child rows 302–305
# Rows 306+ are excluded entirely

# Excel epoch for serial date conversion
EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_iso(serial) -> str | None:
    """Convert an Excel date serial number to ISO date string."""
    if serial is None:
        return None
    try:
        serial = float(serial)
        if serial < 1:
            return None
        dt = EXCEL_EPOCH + timedelta(days=serial)
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError, OverflowError):
        return None


def safe_numeric(val) -> float | None:
    """Coerce a cell value to float or None. Preserve blanks as None."""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("", "#DIV/0!", "#REF!", "#N/A", "#VALUE!", "-"):
            return None
        try:
            return float(val.replace(",", ""))
        except ValueError:
            return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val) -> str | None:
    """Coerce to string or None."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def is_parent_row(row_data: tuple) -> bool:
    """Detect parent row: column A contains a numeric webinar number."""
    val = row_data[0]  # Column A (0-indexed from row tuple)
    if val is None:
        return False
    try:
        n = int(float(val))
        return 100 <= n <= 999  # webinar numbers are 3-digit
    except (ValueError, TypeError):
        return False


def get_child_kind(list_name_or_title: str | None) -> str:
    """Classify child row kind from column F value."""
    if list_name_or_title is None:
        return "list"
    upper = list_name_or_title.strip().upper()
    if upper == "NONJOINERS":
        return "nonjoiners"
    if upper == "NO LIST DATA":
        return "no_list_data"
    return "list"


def extract_metrics(row_data: tuple) -> dict[str, float | None]:
    """Extract metric values from a row."""
    metrics: dict[str, float | None] = {}
    for col_idx, key in METRIC_KEYS.items():
        cell_val = row_data[col_idx - 1] if col_idx - 1 < len(row_data) else None
        metrics[key] = safe_numeric(cell_val)
    return metrics


def extract_child_row(row_data: tuple, row_num: int) -> dict:
    """Build a child row dict."""
    col_f = safe_str(row_data[5]) if len(row_data) > 5 else None
    kind = get_child_kind(col_f)

    # Column B on child rows is notes/free text
    note = safe_str(row_data[1]) if len(row_data) > 1 else None
    # If note looks like a date serial, convert it
    if note and note.replace(".", "").isdigit():
        converted = excel_serial_to_iso(note)
        if converted:
            note = converted

    return {
        "workbookRow": row_num,
        "kind": kind,
        "status": safe_str(row_data[2]) if len(row_data) > 2 else None,
        "note": note,
        "listUrl": safe_str(row_data[3]) if len(row_data) > 3 else None,
        "description": safe_str(row_data[4]) if len(row_data) > 4 else None,
        "listName": col_f if kind == "list" else col_f,
        "sendInfo": safe_str(row_data[6]) if len(row_data) > 6 else None,
        "descLabel": safe_str(row_data[7]) if len(row_data) > 7 else None,
        "titleText": safe_str(row_data[8]) if len(row_data) > 8 else None,
        "createdDate": excel_serial_to_iso(row_data[13]) if len(row_data) > 13 else None,
        "industry": safe_str(row_data[14]) if len(row_data) > 14 else None,
        "employeeRange": safe_str(row_data[15]) if len(row_data) > 15 else None,
        "country": safe_str(row_data[16]) if len(row_data) > 16 else None,
        "metrics": extract_metrics(row_data),
    }


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/export_statistics_workbook.py <path-to-workbook.xlsx>")

    wb_path = Path(sys.argv[1])
    if not wb_path.exists():
        sys.exit(f"File not found: {wb_path}")

    print(f"Loading workbook: {wb_path}")
    wb = load_workbook(str(wb_path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        ws = wb.worksheets[0]

    webinars: list[dict] = []
    current_webinar: dict | None = None

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_num > MAX_SAFE_ROW:
            break

        if is_parent_row(row):
            # Save previous webinar
            if current_webinar is not None:
                webinars.append(current_webinar)

            webinar_number = int(float(row[0]))
            col_b = row[1] if len(row) > 1 else None
            col_f = safe_str(row[5]) if len(row) > 5 else None

            # Column B on parent rows is a date serial
            date_iso = excel_serial_to_iso(col_b)

            # Title from column F
            title = None
            if col_f:
                if col_f.upper() == "TOTAL":
                    title = "TOTAL"
                elif col_f.upper().startswith("TITLE:"):
                    title = col_f[6:].strip()
                else:
                    title = col_f

            current_webinar = {
                "number": webinar_number,
                "date": date_iso,
                "title": title,
                "workbookRow": row_num,
                "rows": [],
            }
        elif current_webinar is not None:
            # Child row — belongs to current webinar
            child = extract_child_row(row, row_num)
            # Skip completely empty rows (no description, no metrics)
            has_data = (
                child["description"]
                or child["listName"]
                or child["kind"] != "list"
                or any(v is not None for v in child["metrics"].values())
            )
            if has_data:
                current_webinar["rows"].append(child)

    # Don't forget the last webinar
    if current_webinar is not None:
        webinars.append(current_webinar)

    wb.close()

    output = {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "source": "workbook_mock",
        "webinars": webinars,
    }

    out_path = Path(__file__).resolve().parent.parent / "api" / "data" / "statistics_workbook_snapshot.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    total_rows = sum(len(w["rows"]) for w in webinars)
    nonjoiners = sum(1 for w in webinars for r in w["rows"] if r["kind"] == "nonjoiners")
    no_list_data = sum(1 for w in webinars for r in w["rows"] if r["kind"] == "no_list_data")

    print(f"Exported {len(webinars)} webinars, {total_rows} child rows")
    print(f"  Nonjoiners: {nonjoiners}, NO LIST DATA: {no_list_data}")
    print(f"  Output: {out_path}")


if __name__ == "__main__":
    main()
